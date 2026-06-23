import openpyxl
from verifier.models import OrgReport, Finding
from verifier import defects

def test_write_defects_xlsx(tmp_path):
    rep = OrgReport(
        org="X", verdict="FAIL",
        summary={"findings_by_code": {"RULE_FLAG_MISMATCH": 1}},
        findings=(Finding("error", "RULE_FLAG_MISMATCH", "msg", "rule",
                          {"teacher": "T", "member": "M", "weekday": "Tue",
                           "start": "15:00:00", "end": "15:30:00",
                           "template_repeat_until": "2026-07-15"}),),
        coverage={})
    out = tmp_path / "defects.xlsx"
    defects.write_defects_xlsx(rep, str(out))
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Defects"}
    d = wb["Defects"]
    hdr = [d.cell(1, j).value for j in range(1, 16)]
    assert hdr == ["#", "Severity", "Issue", "What's wrong", "Teacher", "Student / Group",
                   "Day", "Time", "Occurrences", "Dates affected", "Teacher ID",
                   "Student/Group ID", "Plan ID", "Fix hint", "Status"]
    assert d.cell(2, 2).value == "Error"          # friendly severity
    assert d.cell(2, 3).value == "RULE_FLAG_MISMATCH"
    assert d.cell(2, 4).value                      # plain-English "What's wrong"
    assert d.cell(2, 5).value == "T"               # teacher
    assert d.cell(2, 8).value == "15:00–15:30"     # time range
    assert "Repeat Until" in d.cell(2, 10).value   # dates affected / value
    assert d.cell(2, 14).value                     # fix hint present
    assert d.cell(2, 15).value == "Open"           # status
