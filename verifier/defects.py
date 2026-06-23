"""Export verifier findings as a developer-facing defect report (.xlsx).

One row per defect with instructor/student/day/time, the specific dates/values, and a
fix hint, so a developer can act without rerunning the tool.
"""
from __future__ import annotations
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# What each finding means for the developer + how to fix it.
FIX_HINTS = {
    "RULE_FLAG_MISMATCH": "Raw plan is open-ended (NoEndDateID set) but the template assigned a "
                          "Repeat Until. Remove the end date so the recurrence stays open-ended.",
    "RULE_MISSING_DATE": "The template recurrence does not generate this raw lesson (usually Start "
                         "Date is later than the series' first lesson). Start the recurrence earlier.",
    "RULE_EXTRA_IN_WINDOW": "The template generates lessons within the raw window that raw never had. "
                            "Check the recurrence cadence / day for this series.",
    "MISSING": "A finite-series lesson present in raw is missing from the template. Add it.",
    "EXTRA": "A finite-series lesson is in the template but not in raw. Remove it.",
    "MISSING_SERIES": "This whole raw enrollment has no template counterpart. Confirm whether its "
                      "occurrence statuses (StatusID) should migrate; if so, the enrollment was lost.",
    "EXTRA_SERIES": "The template has a recurring series with no basis in raw. Remove it.",
    "CANCELLED_STILL_LIVE": "This raw lesson was cancelled but the template still generates it live. "
                            "Carry the cancellation as an exception so it is not scheduled.",
    "SUBSTITUTE_DROPPED": "This substituted lesson is absent from the template entirely. Add it "
                          "(taught by the assigned/substitute teacher).",
    "CADENCE_MISMATCH": "The recurrence cadence (weekly/bi-weekly) differs between raw and template.",
    "DOUBLE_COUNT": "A slot has both a normal and an exception occurrence in the template.",
    "EDGE_ORPHAN": "An edge row references a StudentPlanID that does not exist in raw.",
}

# Plain-English description of the problem (for the person doing validation, not the dev).
WHAT = {
    "MISSING_SERIES": "A recurring lesson series from the live system has no match in the "
                      "converted template — it looks dropped.",
    "EXTRA_SERIES": "The converted template has a recurring series that does not exist in the "
                    "live system.",
    "RULE_FLAG_MISMATCH": "The live series is open-ended but the template gave it an end date "
                          "(or vice versa).",
    "RULE_MISSING_DATE": "The template's recurrence skips a lesson that exists in the live system.",
    "RULE_EXTRA_IN_WINDOW": "The template's recurrence creates a lesson the live system never had.",
    "MISSING": "A one-off lesson in the live system is missing from the template.",
    "EXTRA": "A one-off lesson in the template is not in the live system.",
    "CANCELLED_STILL_LIVE": "A lesson cancelled in the live system is still scheduled (live) in "
                            "the template.",
    "SUBSTITUTE_DROPPED": "A substitute lesson from the live system is missing from the template.",
    "CADENCE_MISMATCH": "Weekly vs bi-weekly cadence differs between the live system and template.",
    "DOUBLE_COUNT": "A time slot has both a normal and an exception lesson in the template "
                    "(double-booked).",
    "EDGE_ORPHAN": "An edge-case row points to a plan ID that does not exist in the live data.",
}
_SEVERITY = {"error": "Error", "warn": "Warning"}

_COLUMNS = ["#", "Severity", "Issue", "What's wrong", "Teacher", "Student / Group", "Day",
            "Time", "Occurrences", "Dates affected", "Teacher ID", "Student/Group ID",
            "Plan ID", "Fix hint", "Status"]

def _clock(v):
    """'16:00:00' -> '16:00'."""
    s = str(v or "")
    return s[:5] if len(s) >= 5 and s[2] == ":" else s

def _time_range(d):
    start, end = _clock(d.get("start")), _clock(d.get("end"))
    return f"{start}–{end}" if (start or end) else ""

def _row(i, f):
    d = f.detail
    affected = d.get("dates") or d.get("date") or ""
    if f.code == "RULE_FLAG_MISMATCH":
        affected = f"template Repeat Until = {d.get('template_repeat_until')}"
    member_id = d.get("group_id") or d.get("member_id") or ""
    return [
        i, _SEVERITY.get(f.severity, f.severity), f.code, WHAT.get(f.code, ""),
        d.get("teacher", ""), d.get("member", ""), d.get("weekday", ""), _time_range(d),
        d.get("count", ""), affected, d.get("teacher_id", ""), member_id,
        d.get("student_plan_id", ""), FIX_HINTS.get(f.code, ""), "Open",
    ]

def build_workbook(report):
    wb = openpyxl.Workbook()

    # Summary sheet
    s = wb.active
    s.title = "Summary"
    s["A1"] = "Org"; s["B1"] = report.org
    s["A2"] = "Verdict"; s["B2"] = report.verdict
    s["A3"] = "Total defects"; s["B3"] = len(report.findings)
    s["A5"] = "Issue"; s["B5"] = "Count"
    for c in ("A1", "A2", "A3", "A5", "B5"):
        s[c].font = Font(bold=True)
    r = 6
    for code, n in sorted(report.summary.get("findings_by_code", {}).items(),
                          key=lambda kv: -kv[1]):
        s.cell(r, 1, code); s.cell(r, 2, n); r += 1

    # Defects sheet
    d = wb.create_sheet("Defects")
    hdr_fill = PatternFill("solid", fgColor="1F2937")
    for j, name in enumerate(_COLUMNS, 1):
        c = d.cell(1, j, name)
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hdr_fill
    # errors first, then by code
    findings = sorted(report.findings,
                      key=lambda f: (0 if f.severity == "error" else 1, f.code))
    red = PatternFill("solid", fgColor="FEE2E2")
    amber = PatternFill("solid", fgColor="FEF3C7")
    for i, f in enumerate(findings, 1):
        vals = _row(i, f)
        for j, v in enumerate(vals, 1):
            cell = d.cell(i + 1, j, v)
            if f.severity == "error":
                cell.fill = red
            elif f.severity == "warn":
                cell.fill = amber
            cell.alignment = Alignment(vertical="top", wrap_text=(j in (4, 10, 14)))
    widths = [4, 9, 22, 46, 20, 22, 5, 12, 11, 34, 11, 14, 10, 50, 10]
    for j, w in enumerate(widths, 1):
        d.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    d.freeze_panes = "A2"
    return wb

def write_defects_xlsx(report, out_path):
    build_workbook(report).save(out_path)
    return out_path

def defects_xlsx_bytes(report):
    """The same defect report as raw .xlsx bytes (for embedding/download)."""
    import io
    buf = io.BytesIO()
    build_workbook(report).save(buf)
    return buf.getvalue()
